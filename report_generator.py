from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment


# -----------------------------
# Styles
# -----------------------------
header_fill = PatternFill(
    fill_type="solid",
    start_color="1F4E78",
    end_color="1F4E78"
)

header_font = Font(
    bold=True,
    color="FFFFFF",
    size=12
)

title_font = Font(
    bold=True,
    size=15
)

thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)

center = Alignment(
    horizontal="center",
    vertical="center"
)


def style_header(cell):
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center
    cell.border = thin_border


def style_cell(cell):
    cell.border = thin_border


def auto_width(sheet):
    for column_cells in sheet.columns:

        length = max(
            len(str(cell.value))
            if cell.value else 0
            for cell in column_cells
        )

        sheet.column_dimensions[
            column_cells[0].column_letter
        ].width = min(length + 5, 60)


# =====================================
# EXISTING LOCAL SEO REPORT
# =====================================
def generate_excel_report(
    business_data,
    website_data,
    local_seo_data,
    scores
):

    wb = Workbook()

    # =====================================
    # Sheet 1 - Business Information
    # =====================================

    ws1 = wb.active
    ws1.title = "Business Information"

    ws1["A1"] = "LOCAL SEO ANALYZER REPORT"
    ws1["A1"].font = title_font

    ws1["A3"] = "Field"
    ws1["B3"] = "Value"

    style_header(ws1["A3"])
    style_header(ws1["B3"])

    row = 4

    for key, value in business_data.items():

        ws1.cell(
            row=row,
            column=1
        ).value = key

        ws1.cell(
            row=row,
            column=2
        ).value = value

        style_cell(
            ws1.cell(row=row, column=1)
        )

        style_cell(
            ws1.cell(row=row, column=2)
        )

        row += 1

    auto_width(ws1)

    # =====================================
    # Sheet 2 - SEO Audit
    # =====================================

    ws2 = wb.create_sheet(
        "SEO Audit Results"
    )

    ws2["A1"] = "Website Analysis"
    ws2["A1"].font = title_font

    ws2["A3"] = "Parameter"
    ws2["B3"] = "Result"

    style_header(ws2["A3"])
    style_header(ws2["B3"])

    row = 4

    for key, value in website_data.items():

        ws2.cell(
            row=row,
            column=1
        ).value = key

        ws2.cell(
            row=row,
            column=2
        ).value = value

        style_cell(
            ws2.cell(row=row, column=1)
        )

        style_cell(
            ws2.cell(row=row, column=2)
        )

        row += 1

    row += 2

    ws2.cell(
        row=row,
        column=1
    ).value = "Local SEO Analysis"

    ws2.cell(
        row=row,
        column=1
    ).font = title_font

    row += 2

    ws2.cell(
        row=row,
        column=1
    ).value = "Parameter"

    ws2.cell(
        row=row,
        column=2
    ).value = "Result"

    style_header(
        ws2.cell(row=row, column=1)
    )

    style_header(
        ws2.cell(row=row, column=2)
    )

    row += 1

    for key, value in local_seo_data.items():

        ws2.cell(
            row=row,
            column=1
        ).value = key

        ws2.cell(
            row=row,
            column=2
        ).value = value

        style_cell(
            ws2.cell(row=row, column=1)
        )

        style_cell(
            ws2.cell(row=row, column=2)
        )

        row += 1

    row += 2

    ws2.cell(
        row=row,
        column=1
    ).value = "SEO SCORE"

    ws2.cell(
        row=row,
        column=1
    ).font = title_font

    row += 2

    ws2.cell(
        row=row,
        column=1
    ).value = "Category"

    ws2.cell(
        row=row,
        column=2
    ).value = "Score"

    style_header(
        ws2.cell(row=row, column=1)
    )

    style_header(
        ws2.cell(row=row, column=2)
    )

    row += 1

    for key, value in scores.items():

        ws2.cell(
            row=row,
            column=1
        ).value = key

        ws2.cell(
            row=row,
            column=2
        ).value = value

        style_cell(
            ws2.cell(row=row, column=1)
        )

        style_cell(
            ws2.cell(row=row, column=2)
        )

        row += 1

    auto_width(ws2)

    # =====================================
    # Sheet 3 - Recommendations
    # =====================================

    ws3 = wb.create_sheet(
        "Recommendations"
    )

    ws3["A1"] = "SEO Recommendations"
    ws3["A1"].font = title_font

    ws3["A3"] = "Recommendation"

    style_header(ws3["A3"])

    recommendations = []

    if website_data["HTTPS"] != "Yes":
        recommendations.append(
            "Enable HTTPS to improve website security."
        )

    if website_data["Mobile Friendly"] != "Yes":
        recommendations.append(
            "Make the website mobile friendly."
        )

    if website_data["Meta Title"] == "Not Found":
        recommendations.append(
            "Add a Meta Title."
        )

    if website_data["Meta Description"] == "Not Found":
        recommendations.append(
            "Add a Meta Description."
        )

    if website_data["H1 Tag"] == "Not Found":
        recommendations.append(
            "Add an H1 heading."
        )

    if website_data["Sitemap"] != "Found":
        recommendations.append(
            "Create a sitemap.xml file."
        )

    if website_data["Robots.txt"] != "Found":
        recommendations.append(
            "Create a robots.txt file."
        )

    if website_data["Favicon"] != "Found":
        recommendations.append(
            "Add a favicon."
        )

    if website_data["Contact Information"] != "Found":
        recommendations.append(
            "Display contact information on the website."
        )

    if website_data["Google Maps Embedded"] != "Found":
        recommendations.append(
            "Embed Google Maps on the Contact page."
        )

    if website_data["WhatsApp Button"] != "Found":
        recommendations.append(
            "Add a WhatsApp contact button."
        )

    if local_seo_data["Location in Title"] != "Yes":
        recommendations.append(
            "Include the location keyword in the page title."
        )

    if (
        local_seo_data[
            "Location in Meta Description"
        ] != "Yes"
    ):
        recommendations.append(
            "Include the location keyword in the meta description."
        )

    if local_seo_data["Location in H1"] != "Yes":
        recommendations.append(
            "Include the location keyword in the H1 heading."
        )

    if local_seo_data["Location in Content"] != "Yes":
        recommendations.append(
            "Mention the location keyword in the website content."
        )

    if not recommendations:
        recommendations.append(
            "Excellent! No major SEO issues detected."
        )

    row = 4

    for rec in recommendations:

        ws3.cell(
            row=row,
            column=1
        ).value = rec

        style_cell(
            ws3.cell(row=row, column=1)
        )

        row += 1

    auto_width(ws3)

    wb.save(
        "SEO_Report.xlsx"
    )

    print(
        "\n✅ Excel report generated successfully!"
    )

    print(
        "File Name: SEO_Report.xlsx"
    )


# ==========================================
# NEW GOOGLE BUSINESS COMPARISON REPORT
# ==========================================
def generate_comparison_report(
    business1,
    business2,
    comparison
):

    wb = Workbook()

    ws = wb.active
    ws.title = "GBP Comparison"

    # --------------------------------
    # Title
    # --------------------------------
    ws["A1"] = (
        "GOOGLE BUSINESS PROFILE COMPARISON"
    )

    ws["A1"].font = title_font

    # --------------------------------
    # Table Headers
    # --------------------------------
    ws["A3"] = "Profile Detail"

    ws["B3"] = business1.get(
        "Business Name",
        "Business 1"
    )

    ws["C3"] = business2.get(
        "Business Name",
        "Business 2"
    )

    style_header(ws["A3"])
    style_header(ws["B3"])
    style_header(ws["C3"])

    # --------------------------------
    # Comparison Fields
    # --------------------------------
    fields = [
        "Google Rating",
        "Total Reviews",
        "Business Category",
        "Website",
        "Phone Number",
        "Business Hours",
        "Google Maps Link"
    ]

    row = 4

    for field in fields:

        ws.cell(
            row=row,
            column=1
        ).value = field

        ws.cell(
            row=row,
            column=2
        ).value = business1.get(
            field,
            "N/A"
        )

        ws.cell(
            row=row,
            column=3
        ).value = business2.get(
            field,
            "N/A"
        )

        style_cell(
            ws.cell(row=row, column=1)
        )

        style_cell(
            ws.cell(row=row, column=2)
        )

        style_cell(
            ws.cell(row=row, column=3)
        )

        row += 1

    # =====================================
    # Profile Scores
    # =====================================

    row += 2

    ws.cell(
        row=row,
        column=1
    ).value = "PROFILE SCORES"

    ws.cell(
        row=row,
        column=1
    ).font = title_font

    row += 1

    ws.cell(
        row=row,
        column=1
    ).value = "Business"

    ws.cell(
        row=row,
        column=2
    ).value = "Score"

    style_header(
        ws.cell(row=row, column=1)
    )

    style_header(
        ws.cell(row=row, column=2)
    )

    row += 1

    ws.cell(
        row=row,
        column=1
    ).value = business1.get(
        "Business Name"
    )

    ws.cell(
        row=row,
        column=2
    ).value = comparison[
        "Business 1 Score"
    ]

    style_cell(
        ws.cell(row=row, column=1)
    )

    style_cell(
        ws.cell(row=row, column=2)
    )

    row += 1

    ws.cell(
        row=row,
        column=1
    ).value = business2.get(
        "Business Name"
    )

    ws.cell(
        row=row,
        column=2
    ).value = comparison[
        "Business 2 Score"
    ]

    style_cell(
        ws.cell(row=row, column=1)
    )

    style_cell(
        ws.cell(row=row, column=2)
    )

    # =====================================
    # Comparison Summary
    # =====================================

    row += 3

    ws.cell(
        row=row,
        column=1
    ).value = "COMPARISON SUMMARY"

    ws.cell(
        row=row,
        column=1
    ).font = title_font

    # Better profile
    row += 2

    ws.cell(
        row=row,
        column=1
    ).value = "Better Google Business Profile"

    ws.cell(
        row=row,
        column=2
    ).value = comparison[
        "Better Profile"
    ]

    style_cell(
        ws.cell(row=row, column=1)
    )

    style_cell(
        ws.cell(row=row, column=2)
    )

    # --------------------------------
    # Business 1 Strengths
    # --------------------------------
    row += 2

    ws.cell(
        row=row,
        column=1
    ).value = (
        f"Strengths of "
        f"{business1.get('Business Name')}"
    )

    ws.cell(
        row=row,
        column=1
    ).font = header_font

    ws.cell(
        row=row,
        column=1
    ).fill = header_fill

    for strength in comparison[
        "Business 1 Strengths"
    ]:

        row += 1

        ws.cell(
            row=row,
            column=1
        ).value = strength

        style_cell(
            ws.cell(row=row, column=1)
        )

    # --------------------------------
    # Business 2 Strengths
    # --------------------------------
    row += 2

    ws.cell(
        row=row,
        column=1
    ).value = (
        f"Strengths of "
        f"{business2.get('Business Name')}"
    )

    ws.cell(
        row=row,
        column=1
    ).font = header_font

    ws.cell(
        row=row,
        column=1
    ).fill = header_fill

    for strength in comparison[
        "Business 2 Strengths"
    ]:

        row += 1

        ws.cell(
            row=row,
            column=1
        ).value = strength

        style_cell(
            ws.cell(row=row, column=1)
        )

    # --------------------------------
    # Areas for Business 1 Improvement
    # --------------------------------
    row += 2

    ws.cell(
        row=row,
        column=1
    ).value = (
        f"Areas where "
        f"{business1.get('Business Name')} "
        f"can improve"
    )

    ws.cell(
        row=row,
        column=1
    ).font = header_font

    ws.cell(
        row=row,
        column=1
    ).fill = header_fill

    for improvement in comparison[
        "Business 1 Improvements"
    ]:

        row += 1

        ws.cell(
            row=row,
            column=1
        ).value = improvement

        style_cell(
            ws.cell(row=row, column=1)
        )

    auto_width(ws)

    # --------------------------------
    # Save File
    # --------------------------------
    filename = (
        "Google_Business_Profile_Comparison.xlsx"
    )

    wb.save(
        filename
    )

    print(
        "\n✅ Google Business Profile "
        "comparison report generated!"
    )

    print(
        f"File Name: {filename}"
    )